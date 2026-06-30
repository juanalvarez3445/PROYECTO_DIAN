"""Prueba de la hoja de Estadisticas con graficos."""
from datetime import date

import pandas as pd
from openpyxl import load_workbook

from src.consolidacion.conciliador import conciliar
from src.models import Factura
from src.reporte.excel import generar_reporte


def _facturas():
    return [
        Factura(numero_factura="A1", nit_emisor="900", nombre_emisor="RAPPI S.A.S",
                fecha_emision=date(2024, 1, 5), total=3980, medio_pago="Electrónicos"),
        Factura(numero_factura="A2", nit_emisor="800", nombre_emisor="SODIMAC COLOMBIA",
                fecha_emision=date(2024, 2, 10), total=564900, medio_pago="Electrónicos"),
        Factura(numero_factura="A3", nit_emisor="700", nombre_emisor="TIENDA LOCAL",
                fecha_emision=date(2024, 2, 20), total=15000, medio_pago="Efectivo"),
    ]


def _movs():
    return pd.DataFrame([
        {"fecha": date(2024, 1, 6), "descripcion": "PAGO RAPPI", "valor": 3980.0, "tipo": "debito", "referencia": ""},
    ])


def test_reporte_incluye_hoja_estadisticas_con_graficos(tmp_path):
    facturas = _facturas()
    con = conciliar(facturas, _movs(), tolerancia_monto=100, ventana_dias=5)
    ruta = generar_reporte(facturas, con, ruta_salida=tmp_path / "rep.xlsx")

    wb = load_workbook(ruta)
    assert "Estadisticas" in wb.sheetnames
    assert "Resumen" in wb.sheetnames
    # 4 graficos esperados
    assert len(wb["Estadisticas"]._charts) == 4
