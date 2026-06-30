"""Hoja de 'Estadisticas' con graficos nativos de Excel (openpyxl).

Agrega al libro una hoja con tablas de datos agregados y graficos editables:
  1. Estado de conciliacion (torta): conciliadas vs sin conciliar.
  2. Medio de pago (torta): Electronicos vs Efectivo (segun el informe DIAN).
  3. Top 10 proveedores por monto (barras horizontales).
  4. Gasto por mes (columnas) a lo largo del periodo.

Los graficos son objetos nativos de Excel (no imagenes), por lo que se pueden
editar, re-colorear y copiar dentro de Excel.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from ..consolidacion.conciliador import Conciliacion
from ..models import Factura

_AZUL = "0E7C6B"
_FUENTE_TIT = Font(bold=True, size=13, color="15242E")
_FUENTE_ENC = Font(bold=True, color="FFFFFF", size=11)
_FILL_ENC = PatternFill("solid", fgColor=_AZUL)
_MONEDA = '"$"#,##0'


def _df(facturas: list[Factura]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "proveedor": (f.nombre_emisor or "(sin nombre)").strip(),
                "total": float(f.total or 0),
                "fecha": f.fecha_emision,
                "medio": (f.medio_pago or "No especificado").strip() or "No especificado",
            }
            for f in facturas
        ]
    )


def _escribir_tabla(ws, fila0: int, titulo: str, encabezados: list[str], filas: list[tuple], col_moneda: int | None = None):
    """Escribe titulo + tabla y devuelve (fila_encabezado, fila_primer_dato, fila_ultimo_dato)."""
    ws.cell(row=fila0, column=1, value=titulo).font = _FUENTE_TIT
    fila_enc = fila0 + 1
    for j, h in enumerate(encabezados, start=1):
        c = ws.cell(row=fila_enc, column=j, value=h)
        c.font = _FUENTE_ENC
        c.fill = _FILL_ENC
    for i, fila in enumerate(filas):
        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=fila_enc + 1 + i, column=j, value=val)
            if col_moneda and j == col_moneda:
                c.number_format = _MONEDA
                c.alignment = Alignment(horizontal="right")
    return fila_enc, fila_enc + 1, fila_enc + len(filas)


def agregar_hoja_estadisticas(wb, facturas: list[Factura], conciliacion: Conciliacion) -> None:
    df = _df(facturas)
    ws = wb.create_sheet("Estadisticas")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14

    ws.cell(row=1, column=1, value="Estadísticas del informe de facturas recibidas").font = Font(bold=True, size=15)

    # --- 1) Estado de conciliacion (torta) ---
    enc, ini, fin = _escribir_tabla(
        ws, 3, "Estado de conciliación", ["Estado", "Facturas"],
        [("Conciliadas", conciliacion.conciliadas), ("Sin conciliar", conciliacion.sin_match)],
    )
    pie = PieChart()
    pie.title = "Estado de conciliación"
    pie.add_data(Reference(ws, min_col=2, min_row=enc, max_row=fin), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=ini, max_row=fin))
    pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True
    pie.height, pie.width = 7.5, 13
    ws.add_chart(pie, "E3")

    # --- 2) Medio de pago (torta) ---
    medios = df.groupby("medio").agg(cant=("total", "size"), monto=("total", "sum")).reset_index()
    filas_medio = [(r["medio"], int(r["cant"]), float(r["monto"])) for _, r in medios.iterrows()]
    enc, ini, fin = _escribir_tabla(
        ws, 19, "Medio de pago", ["Medio de pago", "Facturas", "Monto"], filas_medio, col_moneda=3,
    )
    pie2 = PieChart()
    pie2.title = "Facturas por medio de pago"
    pie2.add_data(Reference(ws, min_col=2, min_row=enc, max_row=fin), titles_from_data=True)
    pie2.set_categories(Reference(ws, min_col=1, min_row=ini, max_row=fin))
    pie2.dataLabels = DataLabelList(); pie2.dataLabels.showPercent = True
    pie2.height, pie2.width = 7.5, 13
    ws.add_chart(pie2, "E19")

    # --- 3) Top 10 proveedores por monto (barras) ---
    top = df.groupby("proveedor")["total"].sum().sort_values(ascending=False).head(10)
    filas_top = [(prov[:40], float(monto)) for prov, monto in top.items()]
    enc, ini, fin = _escribir_tabla(
        ws, 35, "Top 10 proveedores por monto", ["Proveedor", "Monto"], filas_top, col_moneda=2,
    )
    bar = BarChart()
    bar.type = "bar"  # barras horizontales
    bar.title = "Top 10 proveedores por monto"
    bar.add_data(Reference(ws, min_col=2, min_row=enc, max_row=fin), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=ini, max_row=fin))
    bar.legend = None
    bar.height, bar.width = 9, 18
    ws.add_chart(bar, "E35")

    # --- 4) Gasto por mes (columnas) ---
    dff = df.dropna(subset=["fecha"]).copy()
    if len(dff):
        dff["mes"] = pd.to_datetime(dff["fecha"], errors="coerce").dt.to_period("M").astype(str)
        por_mes = dff.groupby("mes")["total"].sum().reset_index().sort_values("mes")
        filas_mes = [(r["mes"], float(r["total"])) for _, r in por_mes.iterrows()]
    else:
        filas_mes = []
    enc, ini, fin = _escribir_tabla(
        ws, 54, "Gasto por mes", ["Mes", "Monto"], filas_mes, col_moneda=2,
    )
    if filas_mes:
        col = BarChart()
        col.type = "col"
        col.title = "Gasto por mes"
        col.add_data(Reference(ws, min_col=2, min_row=enc, max_row=fin), titles_from_data=True)
        col.set_categories(Reference(ws, min_col=1, min_row=ini, max_row=fin))
        col.legend = None
        col.height, col.width = 9, 18
        ws.add_chart(col, "E54")

    # Dejar la hoja como segunda (despues de Resumen)
    wb.move_sheet("Estadisticas", -(len(wb.sheetnames) - 2))
