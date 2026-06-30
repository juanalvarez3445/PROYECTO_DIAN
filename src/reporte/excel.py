"""Generacion del reporte Excel (.xlsx) con formato profesional.

Crea un libro con cuatro hojas:
  - 'Resumen'            : indicadores clave (totales, conciliadas, etc.).
  - 'Consolidacion'      : el cruce factura <-> movimiento con su estado (tabla principal).
  - 'Facturas recibidas' : una fila por factura.
  - 'Movimientos banco'  : el extracto bancario normalizado.

Aplica formato: encabezados resaltados, filtros, paneles congelados, formato de
moneda y anchos de columna automaticos.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import SALIDA_DIR
from ..consolidacion.conciliador import Conciliacion
from ..models import Factura
from .graficos import agregar_hoja_estadisticas

# Estilos
_AZUL = "0E7C6B"
_GRIS = "F4F7F9"
_FUENTE_ENC = Font(bold=True, color="FFFFFF", size=11)
_FILL_ENC = PatternFill("solid", fgColor=_AZUL)
_MONEDA = '"$"#,##0'


def _facturas_a_df(facturas: list[Factura]) -> pd.DataFrame:
    filas = [
        {
            "Numero": f.numero_factura,
            "NIT emisor": f.nit_emisor,
            "Proveedor": f.nombre_emisor,
            "Fecha emision": f.fecha_emision,
            "Total": f.total,
            "Moneda": f.moneda,
            "CUFE": f.cufe,
        }
        for f in facturas
    ]
    return pd.DataFrame(filas)


def _consolidacion_a_df(conciliacion: Conciliacion) -> pd.DataFrame:
    filas = []
    for r in conciliacion.resultados:
        f = r.factura
        filas.append(
            {
                "Proveedor": f.nombre_emisor,
                "NIT emisor": f.nit_emisor,
                "Numero factura": f.numero_factura,
                "Fecha factura": f.fecha_emision,
                "Total factura": f.total,
                "Estado": "Conciliado" if r.estado == "conciliado" else "Sin conciliar",
                "Fecha movimiento": r.fecha_movimiento,
                "Descripcion movimiento": r.descripcion_movimiento,
                "Valor movimiento": r.valor_movimiento if r.estado == "conciliado" else None,
                "Diferencia": round(r.diferencia, 2) if r.estado == "conciliado" else None,
                "Confianza %": r.puntaje if r.estado == "conciliado" else None,
            }
        )
    return pd.DataFrame(filas)


def _resumen_df(facturas: list[Factura], conciliacion: Conciliacion) -> pd.DataFrame:
    total_fact = sum(f.total for f in facturas)
    conc = [r for r in conciliacion.resultados if r.estado == "conciliado"]
    total_conc = sum(r.factura.total for r in conc)
    n = len(facturas)
    filas = [
        ("Facturas recibidas", n),
        ("Total facturado", total_fact),
        ("Facturas conciliadas", conciliacion.conciliadas),
        ("Monto conciliado", total_conc),
        ("Facturas sin conciliar", conciliacion.sin_match),
        ("Monto sin conciliar", total_fact - total_conc),
        ("% conciliado", f"{round(100 * conciliacion.conciliadas / n)}%" if n else "0%"),
        ("Movimientos en el extracto", len(conciliacion.movimientos)),
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    return pd.DataFrame(filas, columns=["Indicador", "Valor"])


def _formatear_hoja(ws, df: pd.DataFrame, con_filtros: bool = True) -> None:
    # Encabezados
    for j, col in enumerate(df.columns, start=1):
        celda = ws.cell(row=1, column=j)
        celda.font = _FUENTE_ENC
        celda.fill = _FILL_ENC
        celda.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    if con_filtros and len(df):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    # Formato por columna
    for j, col in enumerate(df.columns, start=1):
        letra = get_column_letter(j)
        es_money = any(k in col.lower() for k in ("total", "valor", "monto", "diferencia"))
        ancho = max(len(str(col)), *(len(str(v)) for v in df[col].fillna("").astype(str))) if len(df) else len(str(col))
        ws.column_dimensions[letra].width = min(max(ancho + 3, 10), 48)
        if es_money:
            for i in range(2, len(df) + 2):
                ws.cell(row=i, column=j).number_format = _MONEDA
                ws.cell(row=i, column=j).alignment = Alignment(horizontal="right")


def generar_reporte(
    facturas: list[Factura],
    conciliacion: Conciliacion,
    ruta_salida: str | Path | None = None,
) -> Path:
    """Genera el .xlsx con formato y devuelve la ruta del archivo creado."""
    if ruta_salida is None:
        marca = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_salida = SALIDA_DIR / f"consolidacion_{marca}.xlsx"
    ruta_salida = Path(ruta_salida)

    df_resumen = _resumen_df(facturas, conciliacion)
    df_consol = _consolidacion_a_df(conciliacion)
    df_facturas = _facturas_a_df(facturas)
    df_movs = conciliacion.movimientos

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df_consol.to_excel(writer, sheet_name="Consolidacion", index=False)
        df_facturas.to_excel(writer, sheet_name="Facturas recibidas", index=False)
        df_movs.to_excel(writer, sheet_name="Movimientos banco", index=False)

        _formatear_hoja(writer.sheets["Resumen"], df_resumen, con_filtros=False)
        writer.sheets["Resumen"].column_dimensions["A"].width = 30
        writer.sheets["Resumen"].column_dimensions["B"].width = 22
        _formatear_hoja(writer.sheets["Consolidacion"], df_consol)
        _formatear_hoja(writer.sheets["Facturas recibidas"], df_facturas)
        if len(df_movs):
            _formatear_hoja(writer.sheets["Movimientos banco"], df_movs)

        # Hoja de graficos/estadisticas (objetos nativos de Excel)
        agregar_hoja_estadisticas(writer.book, facturas, conciliacion)

    return ruta_salida
